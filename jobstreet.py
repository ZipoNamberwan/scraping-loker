import requests
import pandas as pd
import time
from datetime import datetime
from bs4 import BeautifulSoup


class JobStreetScraper:
    """A class for scraping job listings from JobStreet Indonesia"""
    
    def __init__(self):
        """Initialize the scraper with required headers, cookies, and parameters"""
        self.cookies = {
            'sol_id': '5b600d91-3699-4b69-8c83-f33134983b1d',
            '_ga': 'GA1.1.1658404211.1758638743',
            '_tt_enable_cookie': '1',
            '_ttp': '01K5VGK85QX7THBQ97K9HMB7NC_.tt.1',
            '_fbp': 'fb.1.1758638743749.735183043440681538',
            'ajs_anonymous_id': 'ff7afac45048707f79632325386a217f',
            '_gcl_au': '1.1.51445185.1758638744',
            'da_cdt': 'visid_019977099112001741f2652009000506f001906700bd0-sesid_1758722117496-hbvid_5b600d91_3699_4b69_8c83_f33134983b1d-tempAcqSessionId_1758722117501-tempAcqVisitorId_5b600d9136994b698c83f33134983b1d',
            '_hjSessionUser_640499': 'eyJpZCI6IjRlOTEzM2VmLTcyOTMtNWExZi05OGJlLWQzZjE0MWQ2YjE4MiIsImNyZWF0ZWQiOjE3NTg2Mzg3NDQ0MzYsImV4aXN0aW5nIjp0cnVlfQ==',
            '_hjHasCachedUserAttributes': 'true',
            '_clck': '1h5hlup%5E2%5Efzl%5E0%5E2092',
            'da_searchTerm': 'undefined',
            '__gads': 'ID=8ad9a68cb5c684fa:T=1758638767:RT=1758722122:S=ALNI_MY4Ytzuptq1nsD4a1_X-t_eJrXyVA',
            '__gpi': 'UID=0000119a3b97aea1:T=1758638767:RT=1758722122:S=ALNI_MZuTETQTZh_ltRrwlkt6v8KFdkSiQ',
            '__eoi': 'ID=e59624fa2967b995:T=1758638767:RT=1758722122:S=AA-Afjahs8Yjg17hOIYmY_KEHh1a',
            'JobseekerSessionId': 'd3183f3b-a12e-4f65-ae9b-06136107c9b6',
            'JobseekerVisitorId': 'd3183f3b-a12e-4f65-ae9b-06136107c9b6',
            'main': 'V%7C2~P%7Cjobsearch~WH%7CJawa%20Barat~WID%7C2030700~OSF%7Cquick&set=1758722863359/V%7C2~P%7Cjobsearch~OSF%7Cquick&set=1758638939137/V%7C2~P%7Cjobsearch~WID%7C2030916~I%7C6281~OSF%7Cquick&set=1758638937615',
            '_cfuvid': 'Lm.9W4CkDj2z3drHTU5qIT4QD0umAmjNBz_qlFVYMKw-1758722863074-0.0.1.1-604800000',
            'utag_main': 'v_id:019977099112001741f2652009000506f001906700bd0$_sn:2$_se:4%3Bexp-session$_ss:0%3Bexp-session$_st:1758724664320%3Bexp-session$ses_id:1758722117496%3Bexp-session$_pn:1%3Bexp-session$_prevpage:search%20results%3Bexp-1758726464330',
            '_ga_DSKCDC8253': 'GS2.1.s1758722118$o3$g1$t1758722864$j60$l0$h0',
            'ttcsid': '1758722118833::34cNnkGvbbhawpixwlUy.2.1758722864372.0',
            '_uetsid': 'fb5f87f0988b11f09d8ca74108b00878',
            '_uetvid': 'fb5f8c50988b11f0bdd25bb2cc52092b',
            'hubble_temp_acq_session': 'id%3A1758722117501_end%3A1758725592628_sent%3A16',
            '_clsk': 'lxc2ni%5E1758739088620%5E1%5E0%5Ey.clarity.ms%2Fcollect',
            '_dd_s': 'rum=0&expire=1758740039406&logs=0',
            'ttcsid_CR8MQEJC77UC0UOHC250': '1758739139429::_EPMV9w6eKCIOorwHxQo.3.1758739139429.0',
        }

        self.headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9,id;q=0.8',
            'priority': 'u=1, i',
            'referer': 'https://id.jobstreet.com/id/jobs/in-Jawa-Barat',
            'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'seek-request-brand': 'jobstreet',
            'seek-request-country': 'ID',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
            'x-seek-checksum': '12a7c5c7',
            'x-seek-site': 'Chalice',
        }
        
        self.base_url = 'https://id.jobstreet.com/api/jobsearch/v5/search'
        self.base_params = {
            'siteKey': 'ID-Main',
            'sourcesystem': 'houston',
            'eventCaptureSessionId': 'b3425f79-1258-4f2d-8f93-c4c825e782c4',
            'userid': 'b3425f79-1258-4f2d-8f93-c4c825e782c4',
            'userqueryid': 'e273e973b2ff4be76d37749857dfb0c3-9139545',
            'usersessionid': 'b3425f79-1258-4f2d-8f93-c4c825e782c4',
            'where': 'Jawa Barat',
            'pageSize': '32',
            'include': 'seodata,gptTargeting',
            'locale': 'id-ID',
            'solId': '5b600d91-3699-4b69-8c83-f33134983b1d',
            'source': 'SEARCH_ENG',
            'facets': 'distinctTitle'
        }
        
        self.delay = 0.2  # seconds between requests

    
    def set_delay(self, delay):
        """Set the delay between requests"""
        self.delay = delay
    
    def get_delay(self):
        """Get the current delay between requests"""
        return self.delay

    def get_single_page_jobs(self, page):
        """Get jobs from a single page"""
        params = self.base_params.copy()
        params['page'] = str(page)
        
        print(f"Scraping JobStreet page {page}...")
        
        try:
            response = requests.get(
                self.base_url,
                params=params,
                cookies=self.cookies,
                headers=self.headers,
            )
            response.raise_for_status()  # Raise an exception for bad status codes
            
            data = response.json()
            jobs = data.get('data', [])

            print(f"Found {len(jobs)} jobs on page {page}")
            return jobs
            
        except requests.exceptions.RequestException as e:
            print(f"Error on page {page}: {e}")
            return None
        except KeyError as e:
            print(f"Error parsing data on page {page}: {e}")
            return None
        except Exception as e:
            print(f"Unexpected error on page {page}: {e}")
            return None
    
    def get_job_detail(self, job_id):
        """Get detailed job content by scraping the HTML job detail page"""
        job_url = f"https://id.jobstreet.com/id/job/{job_id}"
        
        print(f"  Fetching details for job ID: {job_id}")
        
        try:
            response = requests.get(job_url)
            response.raise_for_status()
            
            # Parse HTML content
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find the job details div
            job_details_div = soup.find('div', {'data-automation': 'jobAdDetails'})
            
            if job_details_div:
                # Look for the inner div with the specified classes
                content_div = job_details_div.find('div', class_='kabgy40 pb3f630')
                
                if content_div:
                    # Extract all text content from the div
                    job_content = content_div.get_text(separator='\n', strip=True)
                    
                    if job_content:
                        print(f"    ✓ Content found: {len(job_content)} characters")
                        return {'content': job_content}
                    else:
                        print(f"    ✗ No content found in specified div")
                        return None
                else:
                    print(f"    ✗ Inner content div not found")
                    return None
            else:
                print(f"    ✗ Job details div not found")
                return None
                
        except requests.exceptions.RequestException as e:
            print(f"    Error fetching HTML for job {job_id}: {e}")
            return None
        except Exception as e:
            print(f"    Unexpected error parsing HTML for job {job_id}: {e}")
            return None
    
    def scrape_all_jobs(self, max_pages=None):
        """Scrape all job listings from all pages and save to Excel
        
        Args:
            max_pages (int, optional): Maximum number of pages to scrape. If None, scrape all pages.
        """
        all_jobs = []
        page = 1
        
        if max_pages:
            print(f"Starting JobStreet job scraping (max {max_pages} pages)...")
        else:
            print("Starting JobStreet job scraping (all pages)...")
        
        while True:
            # Check if we've reached the maximum number of pages
            if max_pages and page > max_pages:
                print(f"Reached maximum pages limit ({max_pages}). Stopping...")
                break
            jobs = self.get_single_page_jobs(page)
            
            # Check if jobs retrieval failed or no jobs found
            if jobs is None:
                break
            
            if not jobs:
                print(f"No more jobs found on page {page}. Stopping...")
                break
            
            # Add jobs to our collection with details
            for job in jobs:
                job_id = job.get('id')
                if job_id:
                    # Fetch job details
                    job_detail = self.get_job_detail(job_id)
                    if job_detail:
                        job['detail'] = job_detail
                    
                    # Add delay between detail requests
                    time.sleep(self.delay)
                
                # Flatten nested data for better Excel representation
                flattened_job = self.flatten_job_data(job)
                all_jobs.append(flattened_job)
            
            # Delay between requests as requested
            time.sleep(self.delay)
            page += 1
        
        # Save to Excel
        if all_jobs:
            self.save_to_excel(all_jobs)
            pages_scraped = min(page-1, max_pages) if max_pages else page-1
            print(f"Successfully scraped {len(all_jobs)} jobs from {pages_scraped} pages")
        else:
            print("No jobs were scraped")
        
        return all_jobs
    
    def scrape_jobs_list_only(self, max_pages=None):
        """Scrape only job listings without details (faster)
        
        Args:
            max_pages (int, optional): Maximum number of pages to scrape. If None, scrape all pages.
        """
        all_jobs = []
        page = 1
        
        if max_pages:
            print(f"Starting JobStreet job scraping (list only, max {max_pages} pages)...")
        else:
            print("Starting JobStreet job scraping (list only, all pages)...")
        
        while True:
            # Check if we've reached the maximum number of pages
            if max_pages and page > max_pages:
                print(f"Reached maximum pages limit ({max_pages}). Stopping...")
                break
            jobs = self.get_single_page_jobs(page)
            
            # Check if jobs retrieval failed or no jobs found
            if jobs is None:
                break
            
            if not jobs:
                print(f"No more jobs found on page {page}. Stopping...")
                break
            
            # Add jobs to our collection without details
            for job in jobs:
                flattened_job = self.flatten_job_data(job)
                all_jobs.append(flattened_job)
            
            # Delay between requests as requested
            time.sleep(self.delay)
            page += 1
        
        # Save to Excel
        if all_jobs:
            self.save_to_excel(all_jobs)
            pages_scraped = min(page-1, max_pages) if max_pages else page-1
            print(f"Successfully scraped {len(all_jobs)} jobs from {pages_scraped} pages (list only)")
        else:
            print("No jobs were scraped")
        
        return all_jobs
    
    def flatten_job_data(self, job):
        """Flatten nested job data for Excel compatibility - focused on specific fields"""
        flattened = {}
        
        # Core fields you requested
        flattened['id'] = job.get('id')
        flattened['title'] = job.get('title')
        flattened['teaser'] = job.get('teaser')
        flattened['salary_label'] = job.get('salaryLabel', '')
        
        # Advertiser information
        advertiser = job.get('advertiser', {})
        flattened['advertiser_id'] = advertiser.get('id', '')
        flattened['advertiser_description'] = advertiser.get('description', '')
        
        # Classifications (main classification and subclassification)
        classifications = job.get('classifications', [])
        if classifications and len(classifications) > 0:
            first_classification = classifications[0]
            if 'classification' in first_classification:
                flattened['classification'] = first_classification['classification'].get('description', '')
                flattened['classification_id'] = first_classification['classification'].get('id', '')
            if 'subclassification' in first_classification:
                flattened['subclassification'] = first_classification['subclassification'].get('description', '')
                flattened['subclassification_id'] = first_classification['subclassification'].get('id', '')
        else:
            flattened['classification'] = ''
            flattened['classification_id'] = ''
            flattened['subclassification'] = ''
            flattened['subclassification_id'] = ''
        
        # Locations
        locations = job.get('locations', [])
        if locations and len(locations) > 0:
            location_labels = []
            country_codes = []
            for loc in locations:
                location_labels.append(loc.get('label', ''))
                country_codes.append(loc.get('countryCode', ''))
            flattened['locations'] = '; '.join(location_labels)
            flattened['country_codes'] = '; '.join(country_codes)
        else:
            flattened['locations'] = ''
            flattened['country_codes'] = ''
        
        # Additional useful fields
        flattened['listing_date'] = job.get('listingDate', '')
        flattened['listing_date_display'] = job.get('listingDateDisplay', '')
        flattened['work_types'] = '; '.join(job.get('workTypes', []))
        
        # Work arrangements
        work_arrangements = job.get('workArrangements', {})
        if work_arrangements and 'data' in work_arrangements:
            arrangement_labels = []
            for wa in work_arrangements['data']:
                if isinstance(wa, dict) and 'label' in wa:
                    if isinstance(wa['label'], dict) and 'text' in wa['label']:
                        arrangement_labels.append(wa['label']['text'])
                    else:
                        arrangement_labels.append(str(wa['label']))
            flattened['work_arrangements'] = '; '.join(arrangement_labels)
        else:
            flattened['work_arrangements'] = ''
        
        # Job detail information (if available)
        detail = job.get('detail')
        if detail and 'content' in detail:
            # New HTML-based content extraction
            flattened['job_content'] = detail['content']
        else:
            flattened['job_content'] = ''
        
        return flattened
    
    def save_to_excel(self, jobs_data):
        """Save jobs data to Excel file"""
        df = pd.DataFrame(jobs_data)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"jobstreet_jobs_{timestamp}.xlsx"
        
        try:
            # Save to Excel with enhanced formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Jobs', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Jobs']
                
                # Import necessary openpyxl styles
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                
                # Format headers
                for col_num in range(1, len(df.columns) + 1):
                    col_letter = get_column_letter(col_num)
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = wrap_alignment
                
                # Auto-adjust column widths and apply text wrapping
                for col_num, column in enumerate(worksheet.columns, 1):
                    col_letter = get_column_letter(col_num)
                    column_name = df.columns[col_num - 1]
                    
                    max_length = 0
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Special handling for job_content column
                    if 'content' in column_name.lower():
                        # Wide column for content without text wrapping
                        worksheet.column_dimensions[col_letter].width = 80
                    else:
                        # Standard column width
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # Apply alternating row colors
                light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for row_num in range(3, len(df) + 2, 2):  # Every other row starting from row 3
                    for col_num in range(1, len(df.columns) + 1):
                        col_letter = get_column_letter(col_num)
                        worksheet[f"{col_letter}{row_num}"].fill = light_fill
            
            print(f"Data saved to {filename}")
            print(f"Columns saved: {list(df.columns)}")
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            # Fallback to CSV
            csv_filename = f"jobstreet_jobs_{timestamp}.csv"
            df.to_csv(csv_filename, index=False)
            print(f"Fallback: Data saved to {csv_filename}")


# Run the scraper
if __name__ == "__main__":
    scraper = JobStreetScraper()
    
    # Choose scraping method:
    # Option 1: Scrape with job details (slower but more comprehensive)
    print("Starting comprehensive JobStreet scraping with job details...")
    scraper.scrape_all_jobs()  # Scrape all pages
    
    # Option 2: Scrape only first 3 pages with details
    # scraper.scrape_all_jobs(max_pages=1)
    
    # Option 3: Scrape only first 5 pages without details (faster)
    # scraper.scrape_jobs_list_only(max_pages=5)
    
    # Option 4: Scrape all pages without details (faster)
    # scraper.scrape_jobs_list_only()


